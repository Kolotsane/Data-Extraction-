import PyPDF2
import os
import re
import openpyxl
import xlsxwriter

# Workbook() takes one, non-optional, argument
# which is the filename that we want to create.
workbook = xlsxwriter.Workbook('Invoice_Information.xlsx')

# The workbook object is then used to add new
# worksheet via the add_worksheet() method.
worksheet = workbook.add_worksheet()

# Use the worksheet object to write
# data via the write() method.
worksheet.write('A1', 'Company Name')
worksheet.write('B1', 'Invoice Number')
worksheet.write('C1', 'Email')
worksheet.write('D1', 'Date')
worksheet.write('E1', 'Total')

workbook.close()
# Opening the created excel file
excelsheet = openpyxl.load_workbook('Invoice_Information.xlsx')
sheet = excelsheet['Sheet1']

# Extracting data from the multiple pdf files
for file_name in os.listdir('invoices'):
    print(file_name)
    file_name = "invoice2.pdf"
    load_pdf = open(r'C:\\Users\\KOLOTSANE\\PycharmProjects\\DataExtraction\\invoices\\' + file_name, 'rb')
    read_pdf = PyPDF2.PdfFileReader(load_pdf, strict=False)
    page_count = read_pdf.getNumPages()
    first_page = read_pdf.getPage(0)
    page_content = first_page.extractText()

    try:

        # print(page_content)
        # Finding the array of totals and storing them in variables
        total = re.findall(r'(?<!\S)(?:(?:cad|[$]|usd|R|M|P) ?[\d,.]+|[\d.,]+(?:cad|[$]|usd))(?!\S)', page_content)
        # if total:
        #     total1 = total[len(total)-1]
        #     print(total1)
        # else:
        #     total1 = " "
        print(total)

        # Finding the array of Invoice dates and storing them in variables
        invoice_date = re.findall(r'(?:\d{1,2}[-/th|st|nd|rd\s]*)?(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)?[a-z\s,.]*(?:\d{1,2}[-/th|st|nd|rd)\s,]*)+(?:\d{2,4})+', page_content)
        print(invoice_date)
        # # sorted(invoice_date, key=lambda x: datetime.datetime.strptime(x, '%d-|/| %m-|/| %Y'))
        # for i in invoice_date:
        #     match = re.search(r"[0-9]{2}\/[0-9]{2}\/[0-9]{4}", i)
        #     match1 = re.search(r"[0-9]{2}\-[0-9]{2}\-[0-9]{4}", i)
        #     match2 = re.findall(r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)+[a-z\s,./]*(?:\d{1,2}[-/th|st|nd|rd)\s,]*)+(?:\d{4})', i)
        #     if match:
        #         date = match.group()
        #         # date1 = date.split('\n', 0)[0]
        #         print(date)
        #     elif match1:
        #         date = match1.group()
        #         print(date)
        #     elif match2:
        #         date = match2[0]
        #         print(date)
        #     # else:
        #     #     date = " "

        # Finding the array of emails and storing them in variables
        email = re.findall(r'([a-zA-Z0-9._-]+@[a-zA-Z0-9._-]+\.[a-zA-Z0-9._-]+)', page_content)
        print(email)
        # if email:
        #     email1 = email[0]
        #     company_name = re.search(r'([\w\.-]+)@([\w\.-]+)', email1).group(2).split(".")[0]
        #     print(email1)
        #     print(company_name)
        # else:
        #     email1 = " "
        #     company_name = " "

        # Finding the array of Invoice Numbers and storing them in variables

        invoice_no = re.findall(r'[a-zA-Z._-]{2,4}[0-9]{4,12}', page_content)
        # invoice_no = re.search(r'Invoice Number|Invoice #|Invoice No.|Invoice Number:|Invoice Number.|Invoice #.|Invoice No:(.*)', page_content).group()
        # if invoice_no:
        #     invoice_no1 = invoice_no[len(invoice_no)-1]
        #     print(invoice_no1)
        # else:
        #     invoice_no1 = " "
        print(invoice_no)

        # Finding the array of company names and storing them in variables
        # company_name = re.findall(r'\b[A-Z]\w+(?:\.com?)?(?:[ -]+(?:&[ -]+)?[A-Z]\w+(?:\.com?)?){0,2}[,\s]+(?i:ltd|llc|inc|plc|co(?:rp)?|group|holding|gmbh)\b', page_content)
        # if company_name:
        #     print(company_name)
        # else:
        #     company_name = " "

        print('\n')
        # Storing data in excel file

        # last_row_number = sheet.max_row
        # print(last_row_number)
        #
        # sheet.cell(column=1, row=last_row_number + 1).value = company_name
        # sheet.cell(column=2, row=last_row_number + 1).value = invoice_no1
        # sheet.cell(column=3, row=last_row_number + 1).value = email1
        # sheet.cell(column=4, row=last_row_number + 1).value = date
        # sheet.cell(column=5, row=last_row_number + 1).value = total1

        # saving a file
        excelsheet.save('Invoice_Information.xlsx')
    except:
        print("Process failed")
        print('\n')

